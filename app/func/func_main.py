import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep


def teste(nome_da_planilha,mes_elemento_selecionado,ano_elemento_selecionado,primeira_execucao):
    
    # Configuração do WebDriver
    service = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=service)

    # Configura o caminho para salvar o arquivo na área de trabalho do usuário
    caminho_area_de_trabalho = os.path.join(os.path.expanduser("~"), "Desktop", f"{nome_da_planilha}.xlsx")

    # Acessa a página
    url = ('https://aplicacoes.mds.gov.br/suaswebcons/restrito/execute.jsf?b=*tbmepQbsdfmbtQbhbtNC&event=*fyjcjs')
    navegador.get(url)

    # Torna os campos que ativam o dropdown de município selecionáveis
    def selecionar_elementos(navegador):
        ano = Select(WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="form:ano"]'))
        ))
        mes = Select(WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="form:mes"]'))
        ))
        esf_adm = Select(WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="form:esferaAdministrativa"]'))
        ))
        uf = Select(WebDriverWait(navegador, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="form:uf"]'))
        ))

        return ano, mes, esf_adm, uf

    # Preenche o valor no campos
    def selecionar_valores_elementos():
        mes.select_by_value(mes_elemento_selecionado)
        sleep(1)
        uf.select_by_visible_text('GO')
        sleep(1)
        ano.select_by_visible_text(ano_elemento_selecionado)
        sleep(1)
        esf_adm.select_by_visible_text('MUNICIPAL')
        sleep(1)

    def click_pesquisar():
        try:
            # Clica no botão pesquisar
            botao_pesquisar = navegador.find_element(By.ID, "form:pesquisar")
            botao_pesquisar.click()

            # Aguarda até que o elemento "valor_elemento" esteja disponível (ou o tempo limite expire)
            WebDriverWait(navegador, 30).until(
                EC.presence_of_element_located((By.XPATH, '//td[@class="tdParNum"]/table/tbody/tr/td[2]/span'))
            )
            print("Página carregada e elemento encontrado.")
        except TimeoutException:
            print("O tempo de espera terminou. O elemento não foi encontrado.")

    while primeira_execucao is True:
        
        navegador.get(url)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Saldo dos Municípios"

        sheet['A1'] = 'MUNICIPIO'
        sheet['B1'] = 'SALDO'
        sheet['C1'] = 'STATUS'

        linha_excel = 2

        ano, mes, esf_adm, uf = selecionar_elementos(navegador)

        selecionar_valores_elementos()
        
        # Aguarda o campo 'municipio' ser ativado (não estar mais 'disabled')
        municipio_dropdown = WebDriverWait(navegador, 20).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#form\\:municipio"))
            )
        municipio = Select(municipio_dropdown)
        

        municipios_valores = [option.get_attribute("value") for option in municipio.options if option.get_attribute("value")]

        for valor in municipios_valores:
            if valor == '522230':
                primeira_execucao = False
            try:
                navegador.refresh()
                ano, mes, esf_adm, uf = selecionar_elementos(navegador)
                
                selecionar_valores_elementos()

                # Aguarda o campo 'municipio' ser ativado novamente
                municipio_dropdown = WebDriverWait(navegador, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#form\\:municipio")))
                municipio = Select(municipio_dropdown)
                municipio.select_by_value(valor)
                selecao_nome_municipio = municipio.first_selected_option.text
                

                click_pesquisar()

                try:
                    # Localiza o segundo <td> dentro da estrutura <td> -> <table> -> <tbody> -> <tr> -> <td>
                    valor_elemento = navegador.find_element(By.XPATH, 
                        '//td[@class="tdParNum"]/table/tbody/tr/td[2]/span')
                    
                    # Extrai o texto visível do elemento
                    valor_saldo = valor_elemento.text.strip()
                    print(f"Valor extraído de {selecao_nome_municipio} é {valor_saldo}")
                    sheet[f'A{linha_excel}'] = selecao_nome_municipio
                    sheet[f'B{linha_excel}'] = valor_saldo
                    sheet[f'C{linha_excel}'] = 'CONCLUIDO'
                    linha_excel += 1



                except Exception as e:
                    sheet[f'A{linha_excel}'] = selecao_nome_municipio
                    sheet[f'C{linha_excel}'] = 'NAO CONCLUIDO'
                    print(f'Deu na captura do valor da linha: {selecao_nome_municipio}')
                    linha_excel += 1
                    
                
                workbook.save(caminho_area_de_trabalho)
                print(f'Dados salvos em {caminho_area_de_trabalho}.')

            except Exception as t:
                sheet[f'A{linha_excel}'] = selecao_nome_municipio
                sheet[f'C{linha_excel}'] = 'NAO CONCLUIDO'
                print(f'Deu erro na seleção dos campos na linha {selecao_nome_municipio}')
                linha_excel += 1
        primeira_execucao = False       

    if primeira_execucao is False:

        workbook = openpyxl.load_workbook(caminho_area_de_trabalho)
        sheet = workbook.active 

        # Itera sobre as linhas da coluna "SALDO" (B), começando pela segunda linha
        # Itera sobre as linhas da coluna "SALDO" (coluna B)

        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=False):  # Apenas a coluna B
            municipio_celula = row[0]
            saldo_celula = row[1]  # Acessa a célula na coluna B (apenas uma célula por linha)
            
            
            if saldo_celula.value is None:
                nome_municipio = municipio_celula.value
                # saldo_celula.value = 'outro'

                try:
                    navegador.get(url)
                    sleep(3)
                    
                    ano, mes, esf_adm, uf = selecionar_elementos(navegador)
                    
                    # Define os valores necessários para ativar o campo de município
                    selecionar_valores_elementos()

                    # Aguarda o campo 'municipio' ser ativado novamente
                    municipio_dropdown = WebDriverWait(navegador, 20).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "#form\\:municipio")))
                    municipio = Select(municipio_dropdown)
                    municipio.select_by_visible_text(nome_municipio)

                    click_pesquisar()

                    try:
                        # Localiza o segundo <td> dentro da estrutura <td> -> <table> -> <tbody> -> <tr> -> <td>
                        valor_elemento = navegador.find_element(By.XPATH, 
                            '//td[@class="tdParNum"]/table/tbody/tr/td[2]/span')
                        
                        # Extrai o texto visível do elemento
                        valor_saldo = valor_elemento.text.strip()
                        saldo_celula.value = valor_saldo
                        print(f"corrigido: {nome_municipio} com o valor: {valor_saldo}")
                        workbook.save(caminho_area_de_trabalho)
                        
                    except Exception as e:
                        print(f'Deu erro na captura do valor da linha: {nome_municipio}')
                        navegador.refresh()

                except Exception as t:
                    print(f'Deu erro na seleção dos campos na linha: {nome_municipio}|info do erro: {t}')
                
    print('ENCERRANDO O PROGRAMA...')
    navegador.quit()